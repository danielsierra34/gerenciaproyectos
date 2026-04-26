from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "INPUT"
OUTPUT = ROOT / "OUTPUT"

SOURCE = INPUT / "Matriz de Pesos.xlsx"
TARGET = OUTPUT / "Matriz_de_Pesos_diligenciada.xlsx"

# Fila: (peso de importancia, Singapur, España, India)
SCORES = {
    8: (4, 4, 5, 3),   # Solidez de la empresa
    9: (5, 4, 4, 3),   # Cumplimiento / referencias
    10: (5, 4, 4, 3),  # Calidad de entregas / referencias
    11: (3, 2, 5, 3),  # Contratos vigentes
    13: (5, 2, 4, 3),  # Experiencia en soluciones similares
    14: (5, 1, 4, 3),  # Experiencia en implantación
    15: (5, 5, 5, 3),  # Experiencia equipo de trabajo
    17: (3, 3, 3, 5),  # Menor cuantía
    18: (4, 4, 5, 4),  # Respaldo financiero
    21: (5, 5, 2, 5),  # Número de CRT
    22: (5, 2, 5, 5),  # Distribución de CRT
    24: (5, 4, 5, 5),  # Data Center
    25: (1, 3, 3, 3),  # RFID
    26: (5, 3, 1, 5),  # Equipos de identificación biométrica
    27: (4, 4, 4, 5),  # Computadores
    28: (5, 3, 4, 5),  # Equipo de comunicaciones
    30: (4, 2, 5, 5),  # Licencias / derechos
    31: (5, 4, 2, 4),  # Requisitos funcionales
    32: (5, 3, 1, 5),  # Compatibilidad biométrica
    33: (5, 4, 5, 5),  # Base de datos alta transaccionalidad
    34: (4, 4, 4, 4),  # Compatibilidad navegadores
    37: (5, 5, 5, 5),  # Seguridad / auditoría
    38: (5, 5, 5, 5),  # Fiabilidad
    39: (5, 5, 5, 5),  # Escalabilidad
    40: (5, 5, 5, 5),  # Rendimiento
    41: (4, 3, 5, 4),  # Usabilidad
    43: (4, 5, 5, 3),  # Configuración / parametrización
    44: (5, 5, 5, 3),  # Recurso humano disponible
    45: (3, 3, 3, 3),  # Compatibilidad con aplicaciones en uso
    46: (4, 4, 4, 4),  # Recursos necesarios para implantación
    47: (4, 3, 4, 3),  # Tiempo necesario
    48: (5, 2, 2, 5),  # Plan de contingencia técnica
}

SECTION_FACTORS = {
    7: (8, 11, 10),
    12: (13, 15, 10),
    20: (21, 22, 4),
    23: (24, 28, 8),
    29: (30, 34, 8),
    36: (37, 41, 15),
    42: (43, 48, 15),
}


def weighted_formula(col, start_row, end_row, factor):
    return f"=SUMPRODUCT($D${start_row}:$D${end_row},{col}{start_row}:{col}{end_row})/SUM($D${start_row}:$D${end_row})/5*{factor}"


def main():
    OUTPUT.mkdir(exist_ok=True)
    wb = load_workbook(SOURCE)
    ws = wb["Plantilla"]
    ws.title = "Matriz de pesos"

    ws["E3"] = "Singapur"
    ws["F3"] = "España"
    ws["G3"] = "India"

    for row, values in SCORES.items():
        for col, value in zip(("D", "E", "F", "G"), values):
            ws[f"{col}{row}"] = value
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Pesos máximos por sección. Las filas detalle conservan la importancia 1-5.
    for cell, value in {
        "D6": 20, "D7": 10, "D12": 10,
        "D16": 30,
        "D19": 20, "D20": 4, "D23": 8, "D29": 8,
        "D35": 30, "D36": 15, "D42": 15,
        "D52": 100,
    }.items():
        ws[cell] = value

    for row, (start_row, end_row, factor) in SECTION_FACTORS.items():
        for col in ("E", "F", "G"):
            ws[f"{col}{row}"] = weighted_formula(col, start_row, end_row, factor)

    for col in ("E", "F", "G"):
        ws[f"{col}6"] = f"={col}7+{col}12"
        ws[f"{col}16"] = weighted_formula(col, 17, 18, 30)
        ws[f"{col}19"] = f"={col}20+{col}23+{col}29"
        ws[f"{col}35"] = f"={col}36+{col}42"
        ws[f"{col}52"] = f"={col}6+{col}16+{col}19+{col}35"

    # Fórmulas equivalentes para dejar clara la estructura de ponderación.
    ws["D52"] = "=D6+D16+D19+D35"

    for row in (6, 16, 19, 35, 52):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            if col >= 4:
                cell.number_format = "0.00"
            if row == 52:
                cell.fill = PatternFill("solid", fgColor="D9EAD3")

    for row in SECTION_FACTORS:
        for col in range(4, 8):
            ws.cell(row=row, column=col).number_format = "0.00"

    ws["C50"] = "Conclusión"
    ws["D50"] = (
        "Con los pesos definidos para un software electoral, India obtiene el mayor puntaje por su menor costo, "
        "cobertura de 12 CRT, distribución nacional, infraestructura AWS, compatibilidad biométrica y plan de contingencia completo. "
        "Antes de adjudicar, se debe mitigar la insuficiencia del equipo y fortalecer el plan de capacitación."
    )
    ws["D50"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("D50:G50")

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).alignment = Alignment(
                horizontal=ws.cell(row=row, column=col).alignment.horizontal or "left",
                vertical="center",
                wrap_text=True,
            )

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 58
    for col in ("D", "E", "F", "G"):
        ws.column_dimensions[col].width = 14

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = "A1:G54"

    wb.save(TARGET)


if __name__ == "__main__":
    main()
