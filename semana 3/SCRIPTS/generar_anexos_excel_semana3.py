from itertools import product
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "OUTPUT"

ACT = {
    "A": {"pred": [], "tn": 1, "cn": 500, "tc": 1, "cc": 500},
    "B": {"pred": ["A"], "tn": 1, "cn": 800, "tc": 1, "cc": 800},
    "C": {"pred": ["B"], "tn": 20, "cn": 1500, "tc": 16, "cc": 2000},
    "D": {"pred": ["B"], "tn": 5, "cn": 150, "tc": 3, "cc": 250},
    "E": {"pred": ["D"], "tn": 2, "cn": 800, "tc": 1, "cc": 1000},
    "F": {"pred": ["B"], "tn": 5, "cn": 150, "tc": 3, "cc": 250},
    "G": {"pred": ["F"], "tn": 10, "cn": 4000, "tc": 10, "cc": 4000},
    "H": {"pred": ["C", "E", "G"], "tn": 1, "cn": 5000, "tc": 1, "cc": 5000},
    "I": {"pred": ["H"], "tn": 5, "cn": 900, "tc": 4, "cc": 1000},
    "J": {"pred": ["B"], "tn": 1, "cn": 100, "tc": 1, "cc": 100},
    "K": {"pred": ["I"], "tn": 1, "cn": 300, "tc": 1, "cc": 300},
    "L": {"pred": ["K"], "tn": 2, "cn": 600, "tc": 2, "cc": 600},
    "M": {"pred": ["K"], "tn": 2, "cn": 400, "tc": 1, "cc": 600},
    "N": {"pred": ["K"], "tn": 1, "cn": 100, "tc": 1, "cc": 100},
    "O": {"pred": ["M", "N"], "tn": 5, "cn": 100, "tc": 3, "cc": 200},
    "P": {"pred": ["L", "O"], "tn": 15, "cn": 1000, "tc": 5, "cc": 1500},
    "Q": {"pred": ["P"], "tn": 1, "cn": 400, "tc": 1, "cc": 400},
}
ORDER = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q"]
INDIRECT_BASE = 7800
SAVING_PER_WEEK = 150


def schedule(dur):
    es, ef = {}, {}
    for a in ORDER:
        es[a] = max([ef[p] for p in ACT[a]["pred"]] or [0])
        ef[a] = es[a] + dur[a]
    proj = max(ef.values())
    succ = {a: [] for a in ORDER}
    for a in ORDER:
        for p in ACT[a]["pred"]:
            succ[p].append(a)
    lf, ls = {}, {}
    for a in reversed(ORDER):
        lf[a] = proj if not succ[a] else min(ls[s] for s in succ[a])
        ls[a] = lf[a] - dur[a]
    slack = {a: ls[a] - es[a] for a in ORDER}
    crit = [a for a in ORDER if abs(slack[a]) < 1e-9]
    return proj, es, ef, ls, lf, slack, crit, succ


def direct_cost(dur):
    total = 0.0
    for a, v in ACT.items():
        if v["tn"] == v["tc"]:
            total += v["cn"]
        else:
            slope = (v["cc"] - v["cn"]) / (v["tn"] - v["tc"])
            total += v["cn"] + slope * (v["tn"] - dur[a])
    return total


def solve_best():
    normal = {a: v["tn"] for a, v in ACT.items()}
    normal_duration, *_ = schedule(normal)
    crashable = [a for a, v in ACT.items() if v["tn"] != v["tc"]]
    ranges = {a: list(range(ACT[a]["tc"], ACT[a]["tn"] + 1)) for a in crashable}
    best = None
    best_by_dur = {}
    for vals in product(*[ranges[a] for a in crashable]):
        dur = {a: v["tn"] for a, v in ACT.items()}
        for a, d in zip(crashable, vals):
            dur[a] = d
        proj, *_ = schedule(dur)
        dcost = direct_cost(dur)
        red = normal_duration - proj
        icost = INDIRECT_BASE - SAVING_PER_WEEK * red
        tcost = dcost + icost
        if best is None or tcost < best["total"]:
            best = {"total": tcost, "duration": proj, "direct": dcost, "indirect": icost, "dur": dur}
        cur = best_by_dur.get(proj)
        if cur is None or tcost < cur["total"]:
            best_by_dur[proj] = {"total": tcost, "direct": dcost, "indirect": icost, "dur": dur}
    return normal_duration, best, best_by_dur


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="D9EAD3")
    for c in ws[row]:
        c.font = Font(bold=True)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 50)


def build_cronograma_excel():
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Datos"
    ws_data.append(["Actividad", "Predecesoras", "Duracion (semanas)"])
    style_header(ws_data, 1)
    for a in ORDER:
        pred = ", ".join(ACT[a]["pred"]) if ACT[a]["pred"] else "INICIO"
        ws_data.append([a, pred, ACT[a]["tn"]])
    autosize(ws_data)

    normal = {a: v["tn"] for a, v in ACT.items()}
    dur, es, ef, ls, lf, slack, crit, succ = schedule(normal)

    ws = wb.create_sheet("Calculo_CPM")
    ws.append(["Actividad", "Predecesoras", "Sucesoras", "Duracion", "ES", "EF", "LS", "LF", "Holgura", "Critica"])
    style_header(ws, 1)
    for a in ORDER:
        pred = ", ".join(ACT[a]["pred"]) if ACT[a]["pred"] else "INICIO"
        suc = ", ".join(succ[a]) if succ[a] else "FIN"
        ws.append([a, pred, suc, ACT[a]["tn"], es[a], ef[a], ls[a], lf[a], slack[a], "Si" if a in crit else "No"])

    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 10).value == "Si":
            for c in range(1, 11):
                ws.cell(r, c).fill = PatternFill("solid", fgColor="FDE9E9")
    autosize(ws)

    ws_r = wb.create_sheet("Resumen")
    ws_r.append(["Indicador", "Valor"])
    style_header(ws_r, 1)
    ws_r.append(["Duracion total (semanas)", dur])
    ws_r.append(["Ruta critica", "-".join(crit)])
    ws_r.append(["Actividades con holgura > 0", ", ".join([a for a in ORDER if slack[a] > 0])])
    ws_r.append(["Nota", "A usa INICIO y Q usa FIN como hitos ficticios para continuidad de red."])
    autosize(ws_r)

    ws_n = wb.create_sheet("Notas_autor")
    ws_n.append(["Notas del autor"])
    style_header(ws_n, 1)
    notes = [
        "Modele la red actividad-nodo porque asi reviso ES/EF/LS/LF sin confusiones.",
        "Trabaje en semanas enteras para mantener consistencia con el enunciado.",
        "No use traslapes: primero quise dejar claro el CPM base.",
        "A y Q usan hitos ficticios INICIO/FIN para continuidad de red.",
        "La ruta critica de este archivo alimenta el analisis de compresion.",
    ]
    for n in notes:
        ws_n.append([n])
    autosize(ws_n)

    wb.save(OUTPUT / "Anexo_Excel_Cronograma_SGNP.xlsx")


def build_compresion_excel():
    normal_duration, best, best_by_dur = solve_best()
    best_dur = best["dur"]
    _, es, ef, ls, lf, slack, crit, succ = schedule(best_dur)

    wb = Workbook()
    ws_d = wb.active
    ws_d.title = "Datos_Compresion"
    ws_d.append(["Actividad", "Predecesoras", "T normal", "C normal", "T compresion", "C compresion", "Costo comp/sem"])
    style_header(ws_d, 1)
    for a in ORDER:
        v = ACT[a]
        slope = 0 if v["tn"] == v["tc"] else (v["cc"] - v["cn"]) / (v["tn"] - v["tc"])
        ws_d.append([a, ", ".join(v["pred"]) if v["pred"] else "INICIO", v["tn"], v["cn"], v["tc"], v["cc"], round(slope, 2)])
    autosize(ws_d)

    ws_i = wb.create_sheet("Iteraciones")
    ws_i.append(["Paso", "Ruta critica activa", "Criterio", "Actividad elegida", "Semanas reducidas", "Costo incremental", "Nueva duracion"])
    style_header(ws_i, 1)
    rows = [
        ["0", "A-B-C-H-I-K-M-O-P-Q", "Base", "Ninguna", 0, 0, 52],
        ["1", "A-B-C-H-I-K-M-O-P-Q", "Menor costo/sem en critica", "P", 10, 500, 42],
        ["2", "A-B-C-H-I-K-M-O-P-Q", "Menor costo/sem en critica", "O", 2, 100, 40],
        ["3", "A-B-C-H-I-K-M-O-P-Q y A-B-C-H-I-K-L-P-Q", "Actividad comun de rutas criticas", "P (comun)", 0, 0, 40],
        ["4", "A-B-C-H-I-K-M-O-P-Q", "Siguiente menor costo/sem", "I", 1, 100, 39],
        ["5", "A-B-C-H-I-K-M-O-P-Q", "Siguiente menor costo/sem", "C", 4, 500, 35],
    ]
    for r in rows:
        ws_i.append(r)
    autosize(ws_i)

    ws_c = wb.create_sheet("Costo_Total")
    ws_c.append(["Duracion (semanas)", "Costo directo", "Costo indirecto", "Costo total"])
    style_header(ws_c, 1)
    for d in sorted(best_by_dur):
        item = best_by_dur[d]
        ws_c.append([d, round(item["direct"], 2), round(item["indirect"], 2), round(item["total"], 2)])
    for r in range(2, ws_c.max_row + 1):
        if ws_c.cell(r, 1).value == best["duration"]:
            for c in range(1, 5):
                ws_c.cell(r, c).fill = PatternFill("solid", fgColor="FDE9E9")
    ws_c.append([])
    ws_c.append(["Duracion optima", best["duration"], "Costo total optimo", round(best["total"], 2)])
    ws_c.append(["Costo directo optimo", round(best["direct"], 2), "Costo indirecto optimo", round(best["indirect"], 2)])
    autosize(ws_c)

    ws_r = wb.create_sheet("Red_Optima")
    ws_r.append(["Actividad", "Predecesoras", "Sucesoras", "Duracion optima", "ES", "EF", "LS", "LF", "Holgura", "Critica"])
    style_header(ws_r, 1)
    for a in ORDER:
        pred = ", ".join(ACT[a]["pred"]) if ACT[a]["pred"] else "INICIO"
        suc = ", ".join(succ[a]) if succ[a] else "FIN"
        ws_r.append([a, pred, suc, best_dur[a], es[a], ef[a], ls[a], lf[a], slack[a], "Si" if a in crit else "No"])
    autosize(ws_r)

    ws_n = wb.create_sheet("Notas_autor")
    ws_n.append(["Notas del autor"])
    style_header(ws_n, 1)
    notes = [
        "Solo comprimi actividades criticas y con margen real de compresion.",
        "Priorice menor costo por semana comprimida.",
        "Con rutas criticas multiples, busque actividades comunes para bajar costo incremental.",
        "Use ahorro indirecto fijo de 150 por semana reducida.",
        "El mejor punto fue 35 semanas por costo total minimo.",
        "Asumi pendiente lineal entre tiempo normal y tiempo de compresion.",
    ]
    for n in notes:
        ws_n.append([n])
    autosize(ws_n)

    wb.save(OUTPUT / "Anexo_Excel_Compresion_SGNP.xlsx")


def main():
    OUTPUT.mkdir(exist_ok=True)
    build_cronograma_excel()
    build_compresion_excel()


if __name__ == "__main__":
    main()
